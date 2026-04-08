"""
Sales Report Automation
Reads a sales CSV, cleans data, summarizes by region/product,
calculates month-over-month growth, and exports a formatted Excel report.
"""

import sys
import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
COMPANY_NAME = "Acme Corp"
REPORT_TITLE = "Sales Performance Report"

HEADER_BG   = "1F3864"   # dark navy
HEADER_FG   = "FFFFFF"
SUBHEAD_BG  = "2E75B6"   # medium blue
SUBHEAD_FG  = "FFFFFF"
ACCENT_BG   = "D6E4F0"   # light blue row shading
TABLE_BORDER = "B8CCE4"

REQUIRED_COLUMNS = {"date", "product", "region", "revenue", "units_sold"}


# ---------------------------------------------------------------------------
# 1. Load & Clean
# ---------------------------------------------------------------------------
def load_and_clean(csv_path: str) -> pd.DataFrame:
    """Load CSV and return a clean DataFrame."""
    path = Path(csv_path)
    if not path.exists():
        sys.exit(f"[ERROR] File not found: {csv_path}")

    df = pd.read_csv(path)

    # Normalise column names
    df.columns = df.columns.str.strip().str.lower().str.replace(r"\s+", "_", regex=True)

    missing = REQUIRED_COLUMNS - set(df.columns)
    if missing:
        sys.exit(f"[ERROR] Missing required columns: {missing}")

    original_rows = len(df)

    # --- date ---
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    bad_dates = df["date"].isna().sum()
    df = df.dropna(subset=["date"])

    # --- numeric ---
    for col in ("revenue", "units_sold"):
        df[col] = pd.to_numeric(
            df[col].astype(str).str.replace(r"[,$€£\s]", "", regex=True),
            errors="coerce",
        )

    # Drop rows missing critical numeric data
    df = df.dropna(subset=["revenue", "units_sold"])

    # Fill missing categorical fields
    df["product"] = df["product"].fillna("Unknown").str.strip()
    df["region"]  = df["region"].fillna("Unknown").str.strip()

    # Derived columns
    df["year_month"] = df["date"].dt.to_period("M")
    df["month_label"] = df["date"].dt.strftime("%b %Y")

    cleaned_rows = len(df)
    dropped = original_rows - cleaned_rows
    print(f"[INFO] Loaded {original_rows} rows → kept {cleaned_rows} "
          f"({dropped} dropped, {bad_dates} bad dates)")
    return df.sort_values("date").reset_index(drop=True)


# ---------------------------------------------------------------------------
# 2. Summaries
# ---------------------------------------------------------------------------
def region_summary(df: pd.DataFrame) -> pd.DataFrame:
    summary = (
        df.groupby("region", as_index=False)
        .agg(total_revenue=("revenue", "sum"), total_units=("units_sold", "sum"))
        .sort_values("total_revenue", ascending=False)
    )
    summary["avg_price"] = (summary["total_revenue"] / summary["total_units"]).round(2)
    summary["revenue_share_%"] = (
        summary["total_revenue"] / summary["total_revenue"].sum() * 100
    ).round(1)
    return summary


def product_summary(df: pd.DataFrame) -> pd.DataFrame:
    summary = (
        df.groupby("product", as_index=False)
        .agg(total_revenue=("revenue", "sum"), total_units=("units_sold", "sum"))
        .sort_values("total_revenue", ascending=False)
    )
    summary["avg_price"] = (summary["total_revenue"] / summary["total_units"]).round(2)
    summary["revenue_share_%"] = (
        summary["total_revenue"] / summary["total_revenue"].sum() * 100
    ).round(1)
    return summary


def region_product_pivot(df: pd.DataFrame) -> pd.DataFrame:
    pivot = df.pivot_table(
        index="region", columns="product", values="revenue",
        aggfunc="sum", fill_value=0,
    )
    pivot["Grand Total"] = pivot.sum(axis=1)
    pivot.loc["Grand Total"] = pivot.sum()
    return pivot.round(2)


def monthly_growth(df: pd.DataFrame) -> pd.DataFrame:
    monthly = (
        df.groupby("year_month", as_index=False)
        .agg(revenue=("revenue", "sum"), units=("units_sold", "sum"))
        .sort_values("year_month")
    )
    monthly["month_label"] = monthly["year_month"].dt.strftime("%b %Y")
    monthly["mom_growth_%"] = monthly["revenue"].pct_change().mul(100).round(1)
    monthly["cum_revenue"] = monthly["revenue"].cumsum().round(2)
    return monthly


# ---------------------------------------------------------------------------
# 3. Excel helpers
# ---------------------------------------------------------------------------
def _thin_border(color=TABLE_BORDER):
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_font(bold=True, size=11, color=HEADER_FG):
    return Font(name="Calibri", bold=bold, size=size, color=color)


def _cell_font(bold=False, size=10, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)


def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)


def write_report_header(ws, company: str, report_date: str, title: str):
    """Write a two-row professional header spanning columns A–H."""
    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")

    c1 = ws["A1"]
    c1.value = company
    c1.font = Font(name="Calibri", bold=True, size=16, color=HEADER_FG)
    c1.fill = _fill(HEADER_BG)
    c1.alignment = Alignment(horizontal="center", vertical="center")

    c2 = ws["A2"]
    c2.value = f"{title}  |  Generated: {report_date}"
    c2.font = Font(name="Calibri", bold=False, size=11, color=SUBHEAD_FG)
    c2.fill = _fill(SUBHEAD_BG)
    c2.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20


def write_dataframe(ws, df: pd.DataFrame, start_row: int, section_title: str = None):
    """Write a DataFrame as a formatted table; return the row after the last data row."""
    col_count = len(df.columns) + (1 if df.index.name else 0)
    end_col_letter = get_column_letter(col_count)

    if section_title:
        title_row = start_row
        ws.merge_cells(f"A{title_row}:{end_col_letter}{title_row}")
        tc = ws[f"A{title_row}"]
        tc.value = section_title
        tc.font = Font(name="Calibri", bold=True, size=12, color=SUBHEAD_FG)
        tc.fill = _fill(SUBHEAD_BG)
        tc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[title_row].height = 18
        start_row += 1

    for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True)):
        ws.append(row)
        excel_row = start_row + r_idx - 1  # dataframe_to_rows starts with header

        for c_idx, cell in enumerate(ws.iter_rows(
                min_row=excel_row, max_row=excel_row,
                min_col=1, max_col=col_count), start=0):
            for cell in ws.iter_rows(
                    min_row=excel_row, max_row=excel_row,
                    min_col=1, max_col=col_count):
                for c in cell:
                    c.border = _thin_border()
                    if r_idx == 1:  # header row
                        c.font = _header_font(size=10)
                        c.fill = _fill(HEADER_BG)
                        c.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        shade = r_idx % 2 == 0
                        c.fill = _fill(ACCENT_BG) if shade else _fill("FFFFFF")
                        c.font = _cell_font(size=10)
                        c.alignment = Alignment(horizontal="right" if c.column > 1 else "left")
            break  # one pass per row

    last_data_row = start_row + len(df)  # header + data rows
    return last_data_row + 2             # gap before next section


def autofit_columns(ws, min_width=8, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


# ---------------------------------------------------------------------------
# 4. Sheet writers
# ---------------------------------------------------------------------------
def build_overview_sheet(wb: Workbook, df: pd.DataFrame, report_date: str):
    ws = wb.create_sheet("Overview")
    write_report_header(ws, COMPANY_NAME, report_date, REPORT_TITLE)

    total_rev   = df["revenue"].sum()
    total_units = df["units_sold"].sum()
    num_months  = df["year_month"].nunique()
    date_range  = f"{df['date'].min().strftime('%d %b %Y')} – {df['date'].max().strftime('%d %b %Y')}"

    kpis = [
        ("Date Range",         date_range),
        ("Total Revenue",      f"${total_rev:,.2f}"),
        ("Total Units Sold",   f"{int(total_units):,}"),
        ("Avg Monthly Revenue",f"${total_rev / max(num_months,1):,.2f}"),
        ("Unique Products",    df["product"].nunique()),
        ("Unique Regions",     df["region"].nunique()),
        ("Total Transactions", len(df)),
    ]

    row = 4
    ws[f"A{row}"].value = "Key Performance Indicators"
    ws[f"A{row}"].font = Font(name="Calibri", bold=True, size=12, color=SUBHEAD_FG)
    ws[f"A{row}"].fill = _fill(SUBHEAD_BG)
    ws.merge_cells(f"A{row}:D{row}")
    ws.row_dimensions[row].height = 18
    row += 1

    for label, value in kpis:
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value)
        lc.font = Font(name="Calibri", bold=True, size=10)
        lc.fill = _fill(ACCENT_BG)
        vc.font = _cell_font(size=10)
        lc.border = vc.border = _thin_border()
        lc.alignment = Alignment(vertical="center")
        vc.alignment = Alignment(horizontal="right", vertical="center")
        row += 1

    autofit_columns(ws)
    ws.sheet_view.showGridLines = False


def build_region_sheet(wb: Workbook, df: pd.DataFrame, report_date: str):
    ws = wb.create_sheet("By Region")
    write_report_header(ws, COMPANY_NAME, report_date, "Revenue & Units by Region")

    reg = region_summary(df)

    # Write table starting at row 4
    next_row = 4
    for r_idx, row in enumerate(dataframe_to_rows(reg, index=False, header=True)):
        excel_row = next_row + r_idx
        for c_idx, value in enumerate(row, start=1):
            c = ws.cell(row=excel_row, column=c_idx, value=value)
            c.border = _thin_border()
            if r_idx == 0:
                c.font = _header_font(size=10)
                c.fill = _fill(HEADER_BG)
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.fill = _fill(ACCENT_BG) if r_idx % 2 == 0 else _fill("FFFFFF")
                c.font = _cell_font(size=10)
                c.alignment = Alignment(horizontal="right" if c_idx > 1 else "left")

    chart_start_row = next_row + len(reg) + 3

    # Bar chart – revenue by region
    chart = BarChart()
    chart.type = "col"
    chart.title = "Total Revenue by Region"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Region"
    chart.style = 10
    chart.width = 18
    chart.height = 12

    data_ref = Reference(ws,
                         min_col=2, max_col=2,
                         min_row=next_row, max_row=next_row + len(reg))
    cats_ref = Reference(ws,
                         min_col=1,
                         min_row=next_row + 1, max_row=next_row + len(reg))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, f"A{chart_start_row}")

    autofit_columns(ws)
    ws.sheet_view.showGridLines = False


def build_product_sheet(wb: Workbook, df: pd.DataFrame, report_date: str):
    ws = wb.create_sheet("By Product")
    write_report_header(ws, COMPANY_NAME, report_date, "Revenue & Units by Product")

    prod = product_summary(df)
    next_row = 4

    for r_idx, row in enumerate(dataframe_to_rows(prod, index=False, header=True)):
        excel_row = next_row + r_idx
        for c_idx, value in enumerate(row, start=1):
            c = ws.cell(row=excel_row, column=c_idx, value=value)
            c.border = _thin_border()
            if r_idx == 0:
                c.font = _header_font(size=10)
                c.fill = _fill(HEADER_BG)
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.fill = _fill(ACCENT_BG) if r_idx % 2 == 0 else _fill("FFFFFF")
                c.font = _cell_font(size=10)
                c.alignment = Alignment(horizontal="right" if c_idx > 1 else "left")

    chart_start_row = next_row + len(prod) + 3

    chart = BarChart()
    chart.type = "bar"       # horizontal bars for products
    chart.title = "Total Revenue by Product"
    chart.y_axis.title = "Product"
    chart.x_axis.title = "Revenue ($)"
    chart.style = 10
    chart.width = 18
    chart.height = 12

    data_ref = Reference(ws,
                         min_col=2, max_col=2,
                         min_row=next_row, max_row=next_row + len(prod))
    cats_ref = Reference(ws,
                         min_col=1,
                         min_row=next_row + 1, max_row=next_row + len(prod))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, f"A{chart_start_row}")

    autofit_columns(ws)
    ws.sheet_view.showGridLines = False


def build_pivot_sheet(wb: Workbook, df: pd.DataFrame, report_date: str):
    ws = wb.create_sheet("Region × Product")
    write_report_header(ws, COMPANY_NAME, report_date, "Revenue Pivot: Region × Product")

    pivot = region_product_pivot(df)
    pivot_reset = pivot.reset_index()

    next_row = 4
    for r_idx, row in enumerate(dataframe_to_rows(pivot_reset, index=False, header=True)):
        excel_row = next_row + r_idx
        for c_idx, value in enumerate(row, start=1):
            c = ws.cell(row=excel_row, column=c_idx, value=value)
            c.border = _thin_border()
            is_grand = (
                (r_idx > 0 and str(row[0]) == "Grand Total") or
                (r_idx == 0 and str(value) == "Grand Total")
            )
            if r_idx == 0 or str(row[0]) == "Grand Total":
                c.font = _header_font(size=10)
                c.fill = _fill(HEADER_BG)
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.fill = _fill(ACCENT_BG) if r_idx % 2 == 0 else _fill("FFFFFF")
                c.font = _cell_font(size=10)
                c.alignment = Alignment(horizontal="right" if c_idx > 1 else "left")

    autofit_columns(ws)
    ws.sheet_view.showGridLines = False


def build_monthly_sheet(wb: Workbook, df: pd.DataFrame, report_date: str):
    ws = wb.create_sheet("Monthly Trend")
    write_report_header(ws, COMPANY_NAME, report_date, "Month-over-Month Revenue Growth")

    monthly = monthly_growth(df)
    display_cols = ["month_label", "revenue", "units", "mom_growth_%", "cum_revenue"]
    monthly_display = monthly[display_cols].copy()
    monthly_display.columns = ["Month", "Revenue ($)", "Units Sold",
                                "MoM Growth (%)", "Cumulative Revenue ($)"]

    next_row = 4
    for r_idx, row in enumerate(dataframe_to_rows(monthly_display, index=False, header=True)):
        excel_row = next_row + r_idx
        for c_idx, value in enumerate(row, start=1):
            c = ws.cell(row=excel_row, column=c_idx, value=value)
            c.border = _thin_border()
            if r_idx == 0:
                c.font = _header_font(size=10)
                c.fill = _fill(HEADER_BG)
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.fill = _fill(ACCENT_BG) if r_idx % 2 == 0 else _fill("FFFFFF")
                c.font = _cell_font(size=10)
                c.alignment = Alignment(horizontal="right" if c_idx > 1 else "left")
                # Colour MoM growth cell
                if c_idx == 4 and value is not None and not (
                    isinstance(value, float) and pd.isna(value)
                ):
                    try:
                        c.font = Font(name="Calibri", size=10,
                                      color="375623" if float(value) >= 0 else "C00000",
                                      bold=True)
                    except (TypeError, ValueError):
                        pass

    chart_start_row = next_row + len(monthly_display) + 3

    # Line chart – revenue trend
    line = LineChart()
    line.title = "Monthly Revenue Trend"
    line.y_axis.title = "Revenue ($)"
    line.x_axis.title = "Month"
    line.style = 10
    line.width = 22
    line.height = 14

    rev_ref = Reference(ws,
                        min_col=2, max_col=2,
                        min_row=next_row, max_row=next_row + len(monthly_display))
    cats_ref = Reference(ws,
                         min_col=1,
                         min_row=next_row + 1, max_row=next_row + len(monthly_display))
    line.add_data(rev_ref, titles_from_data=True)
    line.set_categories(cats_ref)

    # Cumulative revenue as a second series
    cum_ref = Reference(ws,
                        min_col=5, max_col=5,
                        min_row=next_row, max_row=next_row + len(monthly_display))
    line.add_data(cum_ref, titles_from_data=True)

    ws.add_chart(line, f"A{chart_start_row}")

    autofit_columns(ws)
    ws.sheet_view.showGridLines = False


def build_raw_sheet(wb: Workbook, df: pd.DataFrame, report_date: str):
    ws = wb.create_sheet("Raw Data")
    write_report_header(ws, COMPANY_NAME, report_date, "Cleaned Transaction Data")

    export_cols = ["date", "product", "region", "revenue", "units_sold"]
    raw = df[export_cols].copy()
    raw["date"] = raw["date"].dt.strftime("%Y-%m-%d")

    next_row = 4
    for r_idx, row in enumerate(dataframe_to_rows(raw, index=False, header=True)):
        excel_row = next_row + r_idx
        for c_idx, value in enumerate(row, start=1):
            c = ws.cell(row=excel_row, column=c_idx, value=value)
            c.border = _thin_border()
            if r_idx == 0:
                c.font = _header_font(size=10)
                c.fill = _fill(HEADER_BG)
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.fill = _fill(ACCENT_BG) if r_idx % 2 == 0 else _fill("FFFFFF")
                c.font = _cell_font(size=10)
                c.alignment = Alignment(horizontal="right" if c_idx > 3 else "left")

    autofit_columns(ws)
    ws.sheet_view.showGridLines = False


# ---------------------------------------------------------------------------
# 5. Main
# ---------------------------------------------------------------------------
def main():
    global COMPANY_NAME
    parser = argparse.ArgumentParser(description="Sales Report Automation")
    parser.add_argument("csv_file", help="Path to the sales CSV file")
    parser.add_argument(
        "--output", "-o",
        default="sales_report.xlsx",
        help="Output Excel file path (default: sales_report.xlsx)",
    )
    parser.add_argument(
        "--company", "-c",
        default=COMPANY_NAME,
        help=f"Company name for report header (default: {COMPANY_NAME})",
    )
    args = parser.parse_args()

    report_date = datetime.now().strftime("%d %b %Y %H:%M")
    COMPANY_NAME = args.company

    print(f"[INFO] Reading: {args.csv_file}")
    df = load_and_clean(args.csv_file)

    print("[INFO] Building Excel report …")
    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    build_overview_sheet(wb, df, report_date)
    build_region_sheet(wb, df, report_date)
    build_product_sheet(wb, df, report_date)
    build_pivot_sheet(wb, df, report_date)
    build_monthly_sheet(wb, df, report_date)
    build_raw_sheet(wb, df, report_date)

    output_path = Path(args.output)
    wb.save(output_path)
    print(f"[INFO] Report saved → {output_path.resolve()}")
    print(f"[INFO] Sheets: {', '.join(ws.title for ws in wb.worksheets)}")


if __name__ == "__main__":
    main()
